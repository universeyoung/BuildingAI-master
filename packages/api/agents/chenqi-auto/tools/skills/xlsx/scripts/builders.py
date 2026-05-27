#!/usr/bin/env python3
"""
Workflow-level builders for tables and charts.

Why this exists
---------------
style_kit.apply_table_style / apply_chart_style are *post-hoc* polish: they
assume the table or chart already exists. In practice models forget to call
them, or call them with stale arguments after they've already touched ~30
openpyxl axis / series / gridline / tick attributes piecemeal — every one of
which silently defaults to a 2007-era Excel preset. The result is a chart
that lints "OK" but looks like a Python data dump.

build_table / build_chart take over at *creation* time so the right defaults
are baked in from the first call. verify_workbook.py then catches anything
that escaped (per CH010-CH018).

Public API
----------
build_table(ws, data, *, anchor='A1', theme='modern_finance',
            column_formats=None, total_row=False,
            band_rows=True, auto_width=True) -> dict
build_chart(ws, *, kind, data_range, series_orient='cols',
            title=None, anchor='B2', size=(720, 360),
            theme='modern_finance', legend='bottom',
            gridlines='horizontal_dashed', y_axis_min='zero',
            y_number_format=None, bar_gap_width=80,
            line_smooth=False, line_markers=True,
            show_axis_titles=False) -> object

Both return useful handles. Both stamp `_builder_meta` on what they create
so verify_workbook can distinguish builder-managed objects from raw ones.
"""

from __future__ import annotations

import re
from typing import Any, Iterable, Sequence

from style_kit import NUMFMT, PALETTES, THEMES, apply_chart_style, apply_table_style


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_RANGE_RE = re.compile(
    r"""^\s*
        (?:'(?P<sheet1>[^']+)'|(?P<sheet2>[^!]+))!
        \$?(?P<c1>[A-Z]+)\$?(?P<r1>\d+)
        :
        \$?(?P<c2>[A-Z]+)\$?(?P<r2>\d+)
        \s*$""",
    re.VERBOSE,
)


def _col_letter(n: int) -> str:
    """1-based column index → letter (1 → A, 27 → AA)."""
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _col_index(letter: str) -> int:
    """Column letter → 1-based index."""
    n = 0
    for c in letter.upper():
        n = n * 26 + (ord(c) - 64)
    return n


def _parse_range(rng: str) -> dict:
    m = _RANGE_RE.match(rng)
    if not m:
        raise ValueError(
            f"Invalid range '{rng}'. Expected 'Sheet!A1:Z9' (single rectangle)."
        )
    return {
        "sheet": m.group("sheet1") or m.group("sheet2"),
        "c1": _col_index(m.group("c1")),
        "r1": int(m.group("r1")),
        "c2": _col_index(m.group("c2")),
        "r2": int(m.group("r2")),
    }


# Splits a multi-range string on commas that sit OUTSIDE single-quoted sheet
# names. Sheet names with literal commas must be wrapped in '...' (Excel
# convention), so we only split on commas not inside a '...' run.
def _split_ranges(s: str) -> list[str]:
    parts, buf, in_quote = [], [], False
    for ch in s:
        if ch == "'":
            in_quote = not in_quote
            buf.append(ch)
        elif ch == "," and not in_quote:
            parts.append("".join(buf).strip())
            buf = []
        else:
            buf.append(ch)
    tail = "".join(buf).strip()
    if tail:
        parts.append(tail)
    return parts


def _parse_ranges(ranges) -> list[dict]:
    """Normalize `data_range` to a list of rectangle dicts.

    Accepts:
      - 'Sheet!A1:D9'                                 → 1 rectangle
      - 'Sheet!A1:A9, Sheet!J1:K9'                    → 2 rectangles
      - ["Sheet!A1:A9", "Sheet!J1:K9"]                → 2 rectangles

    All rectangles must reference the SAME sheet (Excel chart series can
    technically span sheets, but mixing sheets in one chart is an
    anti-pattern that breaks legend/category alignment).
    """
    if isinstance(ranges, str):
        rects = [_parse_range(r) for r in _split_ranges(ranges)]
    elif isinstance(ranges, (list, tuple)):
        if not ranges:
            raise ValueError("data_range list is empty.")
        rects = [_parse_range(r) for r in ranges]
    else:
        raise TypeError(
            f"data_range must be str or list[str], got {type(ranges).__name__}."
        )
    if not rects:
        raise ValueError("data_range yielded zero rectangles.")
    sheets = {r["sheet"] for r in rects}
    if len(sheets) > 1:
        raise ValueError(
            f"data_range rectangles must all be on the same sheet; got {sorted(sheets)}."
        )
    return rects


def _read_cell_values(ws_lookup, sheet: str, c1: int, r1: int, c2: int, r2: int):
    """Return a 2-D list of *resolved* cell values (not formulas)."""
    ws = ws_lookup[sheet]
    out = []
    for r in range(r1, r2 + 1):
        row = []
        for c in range(c1, c2 + 1):
            row.append(ws.cell(row=r, column=c).value)
        out.append(row)
    return out


def _is_dataframe(obj) -> bool:
    return obj.__class__.__name__ == "DataFrame" and hasattr(obj, "columns")


def _df_to_rows(df) -> list[list]:
    cols = list(df.columns)
    rows = [cols]
    for rec in df.itertuples(index=False, name=None):
        rows.append(list(rec))
    return rows


def _estimate_col_width(values: Iterable[Any]) -> float:
    longest = 0
    for v in values:
        if v is None:
            continue
        s = str(v)
        # Rough char-width: CJK ≈ 2 columns, ASCII ≈ 1.
        w = sum(2 if ord(ch) > 127 else 1 for ch in s)
        longest = max(longest, w)
    return min(max(longest + 2, 10), 50)


# ---------------------------------------------------------------------------
# build_table
# ---------------------------------------------------------------------------
def build_table(
    ws,
    data,
    *,
    anchor: str = "A1",
    title: str | None = None,
    title_height: int = 28,
    theme: str = "modern_finance",
    column_formats: dict[str, str] | None = None,
    total_row: bool = False,
    band_rows: bool = True,
    auto_width: bool = True,
) -> dict:
    """Write a table at `anchor` and apply visual treatment in one call.

    Returns a handle dict with concrete row/column coordinates that can be
    fed to build_chart(data_range=...) — never hand-build A1 references.

    Args:
        data:           pandas.DataFrame OR list[list] (first row = header)
        anchor:         top-left cell, e.g. 'A1' — should normally be 'A1' on
                        a fresh sheet. Avoid leaving blank rows / columns
                        around the table; use `title=` for a banner row
                        instead of inserting blank rows manually.
        title:          if given, writes a banner row immediately ABOVE the
                        header row, merged across the full table width with
                        the theme's header_fill / white text.  This is the
                        right way to add a sheet/section title; do NOT call
                        ws.insert_rows(1) + ws['B1']=... afterwards (that
                        leaves the title in a single un-merged cell).
        title_height:   row height for the banner row (px).
        column_formats: {column_name: NUMFMT key}; columns not listed stay
                        as-is. Unknown keys raise ValueError.
        total_row:      append a SUM row over numeric columns
        band_rows:      alternate-row light fill via apply_table_style
        auto_width:     set column widths from the longest cell content

        Note: pane freezing is decided automatically based on table size —
        ≥15 data rows freezes the header; ≥8 columns AND a text-typed first
        column freezes column A as well. Small tables stay un-frozen.

    Returns:
        {
          'title_row':  int | None,
          'header_row': int,
          'data_rows':  (start, end),    # inclusive
          'total_row':  int | None,
          'columns':    {name: 'A'},
          'sheet':      ws.title,
        }
    """
    if column_formats is None:
        column_formats = {}
    bad = [k for k in column_formats.values() if k not in NUMFMT]
    if bad:
        raise ValueError(
            f"column_formats values must be NUMFMT keys; unknown: {bad}. "
            f"Available: {sorted(NUMFMT)}"
        )

    rows = _df_to_rows(data) if _is_dataframe(data) else [list(r) for r in data]
    if not rows or not rows[0]:
        raise ValueError("build_table: data is empty.")

    anchor_info = _parse_range(f"X!{anchor}:{anchor}")  # reuse parser
    col0 = anchor_info["c1"]
    row0 = anchor_info["r1"]

    header = rows[0]
    body = rows[1:]
    n_cols = len(header)

    # Title banner row (above header).
    title_row_idx = None
    if title:
        from openpyxl.styles import Font, PatternFill, Alignment
        title_row_idx = row0
        # Push the table down by 1 row to make room.
        row0 = row0 + 1
        t = THEMES[theme]
        cell = ws.cell(row=title_row_idx, column=col0)
        cell.value = title
        cell.font = Font(name=t["font_name"], size=14, bold=True,
                         color=t["header_font"])
        cell.fill = PatternFill("solid", start_color=t["header_fill"])
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   indent=1)
        if n_cols > 1:
            from openpyxl.utils import get_column_letter
            ws.merge_cells(
                start_row=title_row_idx, start_column=col0,
                end_row=title_row_idx, end_column=col0 + n_cols - 1,
            )
        ws.row_dimensions[title_row_idx].height = title_height

    # Write header
    for j, name in enumerate(header):
        ws.cell(row=row0, column=col0 + j).value = name

    # Write body + per-column number_format
    fmt_per_col: list[str | None] = [
        NUMFMT[column_formats[c]] if c in column_formats else None
        for c in header
    ]
    for i, body_row in enumerate(body):
        r = row0 + 1 + i
        for j in range(n_cols):
            cell = ws.cell(row=r, column=col0 + j)
            cell.value = body_row[j] if j < len(body_row) else None
            if fmt_per_col[j]:
                cell.number_format = fmt_per_col[j]

    data_start = row0 + 1
    data_end = row0 + len(body)

    # Optional SUM total row
    total_row_idx = None
    if total_row and body:
        total_row_idx = data_end + 1
        for j in range(n_cols):
            col_letter = _col_letter(col0 + j)
            cell = ws.cell(row=total_row_idx, column=col0 + j)
            # Heuristic: numeric column if any body cell is numeric.
            if any(isinstance(b[j], (int, float)) for b in body if j < len(b)):
                cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
                if fmt_per_col[j]:
                    cell.number_format = fmt_per_col[j]
            elif j == 0:
                cell.value = "Total"

    # Auto width
    if auto_width:
        for j in range(n_cols):
            col_letter = _col_letter(col0 + j)
            col_values = [header[j]] + [b[j] if j < len(b) else None for b in body]
            ws.column_dimensions[col_letter].width = _estimate_col_width(col_values)

    # Decide pane freezing automatically. The point of freezing is to keep
    # the header / label column visible while scrolling — pointless for
    # tables that fit on a single screen. Thresholds:
    #   - rows: freeze header once data exceeds ~15 rows (≈ a 100%-zoom screen)
    #   - cols: freeze col A once we have ≥8 columns AND col A is text-typed
    #           (numeric first columns aren't useful as a frozen label)
    n_data_rows = data_end - data_start + 1
    first_col_is_label = bool(body) and isinstance(body[0][0], str)
    needs_row_freeze = n_data_rows >= 15
    needs_col_freeze = n_cols >= 8 and first_col_is_label
    freeze_mode = (
        "both" if needs_row_freeze and needs_col_freeze
        else "rows" if needs_row_freeze
        else "cols" if needs_col_freeze
        else "none"
    )

    # Apply baseline visual treatment
    apply_table_style(
        ws,
        header_row=row0,
        data_rows=(data_start, data_end),
        total_row=total_row_idx,
        theme=theme,
        freeze=freeze_mode,
        band_rows=band_rows,
    )

    return {
        "title_row": title_row_idx,
        "header_row": row0,
        "data_rows": (data_start, data_end),
        "total_row": total_row_idx,
        "columns": {name: _col_letter(col0 + j) for j, name in enumerate(header)},
        "sheet": ws.title,
    }


# ---------------------------------------------------------------------------
# build_chart
# ---------------------------------------------------------------------------
_CHART_KINDS = {
    "clustered_bar":  ("BarChart",  {"type": "col", "grouping": "clustered"}),
    "stacked_bar":    ("BarChart",  {"type": "col", "grouping": "stacked",
                                     "overlap": 100}),
    "line":           ("LineChart", {}),
    "area":           ("AreaChart", {}),
    # combo_bar_line is built specially (see _build_combo)
    "combo_bar_line": (None, None),
}


def build_chart(
    ws,
    *,
    kind: str,
    data_range: str,
    series_orient: str = "cols",       # 'cols' = each column is a series
    title: str | None = None,
    anchor: str = "B2",
    size: tuple[int, int] = (720, 360),
    theme: str = "modern_finance",
    legend: str = "bottom",            # 'bottom' | 'right' | 'none'
    gridlines: str = "horizontal_dashed",  # 'none'|'horizontal_dashed'|'horizontal_solid'
    y_axis_min: Any = "zero",          # 'zero' | 'auto_tight' | <number>
    y_number_format: str | None = None,    # None → infer from kind
    bar_gap_width: int = 80,
    line_smooth: bool = False,
    line_markers: bool = True,
    show_axis_titles: bool = False,
    combo_line_series: int | Sequence[int] | None = None,  # for combo only
):
    """Create a chart with sane defaults baked in. Adds it to `ws`.

    See module docstring for rationale. Returns the chart object so the caller
    can attach extra niceties; you generally won't need to.
    """
    if kind not in _CHART_KINDS:
        raise ValueError(
            f"Unknown chart kind '{kind}'. Choose: {sorted(_CHART_KINDS)}"
        )
    if series_orient not in ("rows", "cols"):
        raise ValueError("series_orient must be 'rows' or 'cols'.")
    if theme not in THEMES:
        raise ValueError(f"Unknown theme '{theme}'.")

    rects = _parse_ranges(data_range)
    wb = ws.parent
    sheet_name = rects[0]["sheet"]
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"data_range references sheet '{sheet_name}' not in workbook."
        )
    src_ws = wb[sheet_name]
    ws_lookup = {s: wb[s] for s in wb.sheetnames}

    # Single-rectangle path: keep historical behaviour (first row/col holds
    # both categories AND series headers).  Multi-rectangle path: rectangle
    # #0 is categories (with optional header), rectangles #1..N are pure
    # series data (each with header on its leading row/col per series_orient).
    if len(rects) == 1:
        cat_labels, series_specs = _resolve_single(
            rects[0], ws_lookup, sheet_name, series_orient,
        )
    else:
        cat_labels, series_specs = _resolve_multi(
            rects, ws_lookup, sheet_name, series_orient,
        )

    if not series_specs:
        raise ValueError("data_range yielded zero series — check orientation.")

    series_values = [spec["values"] for spec in series_specs]

    # ---- Build the chart object -------------------------------------------
    if kind == "combo_bar_line":
        chart = _build_combo(
            src_ws, series_specs, combo_line_series, line_smooth, line_markers,
        )
    else:
        chart = _build_simple(
            kind, src_ws, series_specs,
            line_smooth, line_markers, bar_gap_width,
        )

    # ---- Strict literal categories (fix CH015) ----------------------------
    _set_literal_categories(chart, cat_labels)

    # ---- Title (explicit overlay=False so it never sits over plotArea) ----
    if title:
        _apply_title(chart, title)

    # ---- Axes -------------------------------------------------------------
    _apply_axes(
        chart, kind=kind, y_axis_min=y_axis_min,
        y_number_format=y_number_format,
        series_values=series_values,
        show_axis_titles=show_axis_titles,
    )

    # ---- Gridlines --------------------------------------------------------
    _apply_gridlines(chart, gridlines)

    # ---- Legend + plot_area layout (must know about title presence) ------
    _apply_legend_and_layout(chart, legend, has_title=bool(title))

    # ---- Sizing -----------------------------------------------------------
    chart.width  = max(size[0] / 38.0, 10)   # px → cm-ish (Excel uses cm)
    chart.height = max(size[1] / 38.0,  6)

    # ---- Theme: CJK fonts + palette colors --------------------------------
    apply_chart_style(chart, theme=theme,
                      legend_position=None)  # legend already configured

    # ---- Chart frame: subtle grey border, no rounded corners --------------
    _apply_frame(chart)

    # Mark as builder-managed (verify_workbook keys off this).
    chart._builder_meta = {
        "kind": kind, "theme": theme,
        "y_axis_min": str(y_axis_min),
        "size": size,
    }

    ws.add_chart(chart, anchor)
    return chart


# ---------------------------------------------------------------------------
# Internal helpers — series spec resolution
# ---------------------------------------------------------------------------
#
# Series spec shape (consumed by _build_simple / _build_combo):
#   {
#     "name":        str | None,  # legend label; None → openpyxl auto-names
#     "values":      list,        # resolved cell values for axis/format inference
#     "min_col":     int,         # ref bounds INCLUDING the header cell (so we
#     "max_col":     int,         # can pass titles_from_data=True uniformly)
#     "min_row":     int,
#     "max_row":     int,
#     "from_rows":   bool,        # True when the series runs along a row
#   }


def _resolve_single(rect, ws_lookup, sheet, series_orient):
    """Single-rectangle path: original behaviour. First row/col = both
    category labels AND series headers (Excel-table-style)."""
    grid = _read_cell_values(ws_lookup, sheet,
                             rect["c1"], rect["r1"], rect["c2"], rect["r2"])
    if series_orient == "cols":
        cat_labels = [row[0] for row in grid[1:]]
        specs = []
        for j in range(1, len(grid[0])):
            col = rect["c1"] + j
            specs.append({
                "name": grid[0][j],
                "values": [grid[i][j] for i in range(1, len(grid))],
                "min_col": col, "max_col": col,
                "min_row": rect["r1"], "max_row": rect["r2"],
                "from_rows": False,
            })
    else:  # rows
        cat_labels = grid[0][1:]
        specs = []
        for i in range(1, len(grid)):
            row = rect["r1"] + i
            specs.append({
                "name": grid[i][0],
                "values": list(grid[i][1:]),
                "min_col": rect["c1"], "max_col": rect["c2"],
                "min_row": row, "max_row": row,
                "from_rows": True,
            })
    return cat_labels, specs


def _resolve_multi(rects, ws_lookup, sheet, series_orient):
    """Multi-rectangle path: rect[0] holds categories; rects[1:] hold series.

    Per `series_orient`:
      'cols': each rect is a vertical block; first row of each rect = header.
              Categories come from rect[0]'s first column. Each subsequent
              rect contributes (n_cols) series.
              All rects must have the SAME row span as rect[0].
      'rows': each rect is a horizontal block; first column = header.
              Categories come from rect[0]'s first row. Each subsequent rect
              contributes (n_rows) series.
              All rects must have the SAME column span as rect[0].
    """
    cat_rect = rects[0]
    cat_grid = _read_cell_values(ws_lookup, sheet,
                                 cat_rect["c1"], cat_rect["r1"],
                                 cat_rect["c2"], cat_rect["r2"])

    if series_orient == "cols":
        # cats = first column of cat_rect, skipping its header row.
        cat_labels = [row[0] for row in cat_grid[1:]]
        cat_span = (cat_rect["r1"], cat_rect["r2"])  # rows
        for r in rects[1:]:
            if (r["r1"], r["r2"]) != cat_span:
                raise ValueError(
                    "Multi-rectangle data_range: with series_orient='cols' all "
                    f"rectangles must share the same row span. Categories rect "
                    f"spans rows {cat_span[0]}..{cat_span[1]}, but a series rect "
                    f"spans {r['r1']}..{r['r2']}."
                )
        specs = []
        for r in rects[1:]:
            grid = _read_cell_values(ws_lookup, sheet,
                                     r["c1"], r["r1"], r["c2"], r["r2"])
            for j in range(0, len(grid[0])):
                col = r["c1"] + j
                specs.append({
                    "name": grid[0][j],
                    "values": [grid[i][j] for i in range(1, len(grid))],
                    "min_col": col, "max_col": col,
                    "min_row": r["r1"], "max_row": r["r2"],
                    "from_rows": False,
                })
    else:  # rows
        cat_labels = cat_grid[0][1:]
        cat_span = (cat_rect["c1"], cat_rect["c2"])  # cols
        for r in rects[1:]:
            if (r["c1"], r["c2"]) != cat_span:
                raise ValueError(
                    "Multi-rectangle data_range: with series_orient='rows' all "
                    f"rectangles must share the same column span. Categories "
                    f"rect spans cols {cat_span[0]}..{cat_span[1]}, but a "
                    f"series rect spans {r['c1']}..{r['c2']}."
                )
        specs = []
        for r in rects[1:]:
            grid = _read_cell_values(ws_lookup, sheet,
                                     r["c1"], r["r1"], r["c2"], r["r2"])
            for i in range(0, len(grid)):
                row = r["r1"] + i
                specs.append({
                    "name": grid[i][0],
                    "values": list(grid[i][1:]),
                    "min_col": r["c1"], "max_col": r["c2"],
                    "min_row": row, "max_row": row,
                    "from_rows": True,
                })
    return cat_labels, specs


# ---------------------------------------------------------------------------
# Internal helpers — chart construction
# ---------------------------------------------------------------------------
def _build_simple(
    kind, src_ws, series_specs,
    line_smooth, line_markers, bar_gap_width,
):
    from openpyxl.chart import BarChart, LineChart, AreaChart, Reference

    cls_name, attrs = _CHART_KINDS[kind]
    cls_map = {"BarChart": BarChart, "LineChart": LineChart, "AreaChart": AreaChart}
    chart = cls_map[cls_name]()
    for k, v in attrs.items():
        setattr(chart, k, v)

    # Force-disable per-point colour variation. Without an explicit
    # <c:varyColors val="0"/> Excel applies an undocumented "friendly default"
    # for single-series LineChart / AreaChart / BarChart and renders every
    # data point in a different palette colour, with the legend listing the
    # category labels instead of the series name. See SKILL.md note on
    # Choosing `kind` for the failure mode this prevents.
    chart.varyColors = False

    if hasattr(chart, "gapWidth") and "Bar" in cls_name:
        chart.gapWidth = bar_gap_width

    # Use Reference for *values* (so live links work in Excel). Each spec's
    # ref INCLUDES its header cell so we can pass titles_from_data=True
    # uniformly across single- and multi-rectangle paths.
    for spec in series_specs:
        ref = Reference(
            src_ws,
            min_col=spec["min_col"], max_col=spec["max_col"],
            min_row=spec["min_row"], max_row=spec["max_row"],
        )
        chart.add_data(ref, titles_from_data=True, from_rows=spec["from_rows"])

    # Line markers
    if cls_name == "LineChart":
        from openpyxl.chart.marker import Marker
        for s in chart.series:
            s.smooth = line_smooth
            if line_markers:
                s.marker = Marker(symbol="circle", size=6)

    return chart


def _build_combo(
    src_ws, series_specs, combo_line_series, line_smooth, line_markers,
):
    """Build a BarChart + LineChart overlay. Bars use primary y, lines secondary."""
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.marker import Marker

    n = len(series_specs)
    if combo_line_series is None:
        # Default heuristic: last series = line.
        line_idx = {n - 1}
    elif isinstance(combo_line_series, int):
        line_idx = {combo_line_series}
    else:
        line_idx = set(combo_line_series)

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.varyColors = False
    line = LineChart()
    line.varyColors = False

    # Build references series-by-series so we can route them between primary
    # bars and secondary line.
    for i, spec in enumerate(series_specs):
        val_ref = Reference(
            src_ws,
            min_col=spec["min_col"], max_col=spec["max_col"],
            min_row=spec["min_row"], max_row=spec["max_row"],
        )
        target = line if i in line_idx else bar
        target.add_data(val_ref, titles_from_data=True,
                        from_rows=spec["from_rows"])

    for s in line.series:
        s.smooth = line_smooth
        if line_markers:
            s.marker = Marker(symbol="circle", size=6)

    # Combine: line goes on secondary axis.
    line.y_axis.axId = 200
    line.y_axis.crosses = "max"
    bar += line  # openpyxl 3.x supports chart += chart for combos
    return bar


def _set_literal_categories(chart, cat_labels):
    """Replace each series.cat with an inline strLit literal source.

    Why: when cats come from a numRef pointing at TEXT cells (e.g. quarter
    labels like '2024 Q2'), Excel writes an empty <numCache> and the X axis
    renders blank in macOS Excel / WPS / Numbers / Sheet.js previews. Inline
    strLit is independent of any sheet calc engine.
    """
    from openpyxl.chart.data_source import (
        AxDataSource, StrData, StrVal,
    )

    labels = [str(v) if v is not None else "" for v in cat_labels]
    sd = StrData(ptCount=len(labels),
                 pt=[StrVal(idx=i, v=v) for i, v in enumerate(labels)])
    cat_src = AxDataSource(strLit=sd)
    for s in chart.series:
        s.cat = cat_src
    # For combos (BarChart with .series) the .series above only covers bar
    # series; loop overlay charts too if present. Skip self-reference (openpyxl
    # adds chart itself as _charts[0] on reload).
    for sub in getattr(chart, "_charts", []) or []:
        if sub is chart:
            continue
        for s in getattr(sub, "series", []) or []:
            s.cat = cat_src


def _apply_axes(chart, *, kind, y_axis_min, y_number_format,
                series_values, show_axis_titles):
    """Set axis numFmt, tick label position, and y-axis scaling."""
    from openpyxl.chart.axis import ChartLines
    from openpyxl.chart.layout import Layout

    # X axis
    if hasattr(chart, "x_axis") and chart.x_axis is not None:
        chart.x_axis.tickLblPos = "nextTo"
        chart.x_axis.majorTickMark = "out"
        chart.x_axis.delete = False
        if not show_axis_titles:
            chart.x_axis.title = None

    # Y axis (primary)
    if hasattr(chart, "y_axis") and chart.y_axis is not None:
        ax = chart.y_axis
        ax.tickLblPos = "nextTo"
        ax.majorTickMark = "out"
        ax.delete = False
        if not show_axis_titles:
            ax.title = None

        # numFmt: explicit > inferred from kind
        fmt = y_number_format
        if fmt is None:
            # Infer: if all values look like fractions in [-1, 1], use pct;
            # otherwise use thousand-separated int.
            flat = [v for row in series_values for v in row
                    if isinstance(v, (int, float))]
            if flat and all(-1 <= v <= 1 for v in flat) and any(0 < abs(v) < 1 for v in flat):
                fmt = NUMFMT["pct_1"]
            else:
                fmt = NUMFMT["int"]
        ax.number_format = fmt
        # Some openpyxl versions need numFmt set via the descriptor:
        try:
            from openpyxl.chart.axis import NumFmt
            ax.numFmt = NumFmt(formatCode=fmt, sourceLinked=False)
        except Exception:
            pass

        # Scaling (zero / auto_tight / explicit)
        _apply_y_scaling(ax, y_axis_min, series_values)

    # Secondary y axis (combo) — mirror the look
    for sub in getattr(chart, "_charts", []) or []:
        if sub is chart:
            continue
        sub_ax = getattr(sub, "y_axis", None)
        if sub_ax is not None:
            sub_ax.tickLblPos = "nextTo"
            sub_ax.majorTickMark = "out"
            sub_ax.delete = False
            if not show_axis_titles:
                sub_ax.title = None


def _apply_y_scaling(ax, y_axis_min, series_values):
    """Configure y_axis scaling.min/max and majorUnit per mode."""
    flat = [v for row in series_values for v in row
            if isinstance(v, (int, float))]
    if not flat:
        return

    lo, hi = min(flat), max(flat)

    if y_axis_min == "zero" or (isinstance(y_axis_min, (int, float)) and y_axis_min == 0):
        scale_min = 0
        scale_max = hi * 1.1 if hi > 0 else 0
    elif y_axis_min == "auto_tight":
        span = max(hi - lo, abs(hi) * 0.1, 1e-9)
        scale_min = lo - 0.1 * span
        scale_max = hi + 0.1 * span
    else:  # explicit number
        scale_min = float(y_axis_min)
        scale_max = hi * 1.1 if hi > scale_min else scale_min + 1

    if scale_max <= scale_min:
        scale_max = scale_min + 1

    major = _nice_step((scale_max - scale_min) / 5)

    try:
        ax.scaling.min = float(scale_min)
        ax.scaling.max = float(scale_max)
        ax.majorUnit = float(major)
    except Exception:
        pass


def _nice_step(raw: float) -> float:
    """Round a raw step size to a 'nice' 1/2/2.5/5 × 10^k."""
    import math
    if raw <= 0:
        return 1
    exp = math.floor(math.log10(raw))
    base = 10 ** exp
    frac = raw / base
    for nice in (1, 2, 2.5, 5, 10):
        if frac <= nice:
            return nice * base
    return 10 * base


def _apply_gridlines(chart, mode: str):
    from openpyxl.chart.axis import ChartLines
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties

    if not hasattr(chart, "y_axis") or chart.y_axis is None:
        return

    if mode == "none":
        chart.y_axis.majorGridlines = None
    else:
        dash = "dash" if mode == "horizontal_dashed" else None
        ln = LineProperties(w=9525,  # ~0.75pt in EMU
                            solidFill="CCCCCC",
                            prstDash=dash)
        chart.y_axis.majorGridlines = ChartLines(
            spPr=GraphicalProperties(ln=ln)
        )
    # Hide x-axis gridlines for clean look.
    if hasattr(chart, "x_axis") and chart.x_axis is not None:
        chart.x_axis.majorGridlines = None

    # Combo charts: openpyxl initialises overlay LineChart's y_axis with a
    # non-None ChartLines, which serialises to <majorGridlines/> (empty) and
    # Excel renders that as a HEAVY BLACK SOLID grid on top of our dashed
    # primary grid — looks like a horizontal solid line near the bottom.
    # Force secondary-axis gridlines OFF; primary axis carries the grid.
    for sub in getattr(chart, "_charts", []) or []:
        if sub is chart:
            continue
        sub_y = getattr(sub, "y_axis", None)
        if sub_y is not None:
            sub_y.majorGridlines = None
        sub_x = getattr(sub, "x_axis", None)
        if sub_x is not None:
            sub_x.majorGridlines = None


def _apply_title(chart, title: str):
    """Install a title that sits ABOVE the plot area, never overlaying it.

    openpyxl's `chart.title = "string"` shortcut emits a <title> node WITHOUT
    a child <overlay> element. Per OOXML the absence of <overlay> is
    implementation-defined; macOS Excel and Numbers render it as overlaid
    (title floats on top of plotArea). We must construct an explicit Title
    object with overlay=False, and reserve top space in plot_area layout
    (handled by _apply_legend_and_layout).
    """
    from openpyxl.chart.title import Title
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import (
        Paragraph, ParagraphProperties, CharacterProperties,
        RichTextProperties, RegularTextRun,
    )

    cp = CharacterProperties(sz=1400, b=True)
    pp = ParagraphProperties(defRPr=cp)
    para = Paragraph(pPr=pp, r=[RegularTextRun(rPr=cp, t=title)])
    rich = RichText(bodyPr=RichTextProperties(), p=[para])

    t_obj = Title()
    t_obj.tx = openpyxl_chart_text_with_rich(rich)
    t_obj.overlay = False
    chart.title = t_obj


def openpyxl_chart_text_with_rich(rich):
    """Wrap a RichText into the title.tx (ChartContainer/Text) descriptor."""
    from openpyxl.chart.text import Text
    return Text(rich=rich)


def _apply_legend_and_layout(chart, position: str, *, has_title: bool):
    """Configure legend position and overlay flag.

    We intentionally do NOT set manual layout for either the plot area or
    the legend. Excel's automatic layout is dramatically better than any
    fixed rectangle we could pick, because it accounts for:
      - actual y-axis tick label widths (e.g. '18.3%' is wider than '300')
      - legend text wrap (4 long series names → 2 rows; 4 short → 1 row)
      - title height (depends on font + line wrap)
      - secondary y-axis on combo charts (right-side label width)

    Our previous hard-coded `_PLOT_LAYOUTS` table assumed all of these were
    constant across charts, which caused two visible bugs:
      - Long y-axis labels (e.g. '(1.7%)') overran the left frame.
      - 2-row legends overlapped the plot because the manual layout
        clipped legend height to one row.

    The `<overlay val="0"/>` flag on title and legend is sufficient to
    guarantee they don't sit on top of the plot area — Excel auto-positions
    them in the reserved space.
    """
    from openpyxl.chart.legend import Legend

    if position == "none":
        chart.legend = None
        return
    if chart.legend is None:
        chart.legend = Legend()
    code = {"bottom": "b", "right": "r"}.get(position, "b")
    chart.legend.position = code
    chart.legend.overlay = False
    # No legend.layout — let Excel auto-place. A manual layout here forces
    # legend text to wrap into a fixed-height box; if the wrap produces more
    # rows than the box has room for, those extra rows render INSIDE the
    # plot area.


def _apply_frame(chart):
    """Apply a 0.75pt grey square border to chartSpace and ensure the chart
    has a white background.

    Two non-obvious bits:
    1. `chart.roundedCorners = False` is REQUIRED. openpyxl does not emit
       <roundedCorners> by default, and Excel's default is val="1" — so the
       chart frame renders with rounded corners (PowerPoint-clipart look)
       even when we set our own <spPr>. Forcing False writes the explicit
       <roundedCorners val="0"/> tag.
    2. Border weight: 0.25pt is too faint to read as a deliberate frame on
       most displays; 0.75pt (9525 EMU) matches the visual weight of cell
       gridlines around the chart and looks intentional.
    """
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties

    chart.roundedCorners = False
    chart.graphical_properties = GraphicalProperties(
        ln=LineProperties(
            w=9525,             # 0.75 pt (12700 EMU = 1pt)
            solidFill="C8CDD3", # neutral grey, slightly darker than gridlines
        ),
        solidFill="FFFFFF",     # white background, clean against sheet banding
    )


__all__ = ["build_table", "build_chart"]
