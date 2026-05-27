#!/usr/bin/env python3
"""
Structural lint pass for an .xlsx file. Catches problems that recalc cannot.

What it checks
--------------
ERROR (block delivery):
  - CH001  Chart with zero series
  - CH003  Per-point series antipattern (every series val ref is single cell)
  - CH004  Chart with NO category axis
  - CH013  chart.style is set to an Excel built-in (1-48); fight the look back
  - CH015  cat is numRef pointing at TEXT cells / unpopulated numCache
           → X axis renders blank in macOS Excel / WPS / preview tools

WARNING (review before delivery):
  - CH002  > 8 series
  - CH005  Series ref points to non-existent sheet
  - CH010  Gridlines unstyled or dark (default heavy navy solid)
  - CH011  Axis title set while legend is at bottom (overlap risk)
  - CH012  Axis numFmt is General/empty though source column is $ or %
  - CH014  Charts on the same sheet differ by > 30% in size
  - CH016  Y axis missing majorUnit (raw 'naked' axis)
  - CH017  Y axis pinned at 0 but data sits in the top 1/5 (poor visual range)
  - CH018  Bottom legend without overlay=False or explicit layout (overlap)
  - WS001  Wide sheet (>= 8 cols, >= 20 rows) without freeze_panes
  - WS002  Header row appears unstyled
  - WB001  Workbook default font is not CJK-capable

INFO (advisory only):
  - FMT001 number_format not in style_kit.NUMFMT (fragmentation risk)

Implementation note: every check appends to `issues`; nothing prints, raises,
or short-circuits mid-scan. The full report is built once and returned at the
end so a single run surfaces every problem the model needs to fix.

Output JSON shape (mirrors recalc.py for tooling parity):
    {
      "status": "success" | "issues_found" | "error",
      "summary": {"errors": int, "warnings": int, "infos": int},
      "issues": [
        {"severity": "error"|"warning"|"info",
         "code": str,
         "location": str,
         "message": str}, ...
      ]
    }

Usage
-----
    python scripts/verify_workbook.py <xlsx_path>
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

# Single-cell ref like "Sheet!A1" (NO range). Used to flag per-point series.
_SINGLE_CELL_RE = re.compile(
    r"^\s*(?:'[^']+'|[^!]+)!\s*\$?[A-Z]+\$?\d+\s*$"
)

# CJK-capable font families we trust to render Chinese / Japanese / Korean.
_CJK_FONTS = {
    "microsoft yahei", "yahei", "simsun", "simhei", "noto sans cjk sc",
    "noto sans cjk tc", "noto sans cjk jp", "noto sans cjk kr",
    "pingfang sc", "pingfang tc", "hiragino sans gb", "hiragino sans",
    "source han sans sc", "source han sans cn", "source han sans tw",
    "source han sans jp", "source han sans kr", "msyh", "wenquanyi micro hei",
    "苹方", "宋体", "黑体", "微软雅黑",
}

# Known-good number_format strings (case-insensitive). Kept loosely in sync with
# style_kit.NUMFMT. Includes both the "full" forms (with negative + zero rules)
# and common "simple" forms that are widely acceptable.
_KNOWN_NUMFMTS = {
    s.lower() for s in {
        "General", "@", "0", "0.0", "0.00",
        "#,##0", "#,##0.0", "#,##0.00",
        "$#,##0", "$#,##0.00",
        "0%", "0.0%", "0.00%",
        '#,##0;[Red](#,##0);"-"',
        '#,##0.0;[Red](#,##0.0);"-"',
        '#,##0.00;[Red](#,##0.00);"-"',
        '$#,##0;[Red]($#,##0);"-"',
        '$#,##0.00;[Red]($#,##0.00);"-"',
        '0.0%;[Red](0.0%);"-"',
        '0.00%;[Red](0.00%);"-"',
        '0.0"x"',
        "yyyy-mm-dd", '"Q"q yyyy', "mmm yyyy", "mm/dd/yyyy", "yyyy/mm/dd",
    }
}


def _has_text_content(value) -> bool:
    return value is not None and str(value).strip() != ""


def _looks_unstyled_header(cells) -> bool:
    """Heuristic: header row is "naked" if no cell is bold AND no cell has fill."""
    any_bold = False
    any_fill = False
    has_text = False
    for c in cells:
        if not _has_text_content(c.value):
            continue
        has_text = True
        if c.font and c.font.bold:
            any_bold = True
        fill = c.fill
        if fill and getattr(fill, "fgColor", None):
            rgb = getattr(fill.fgColor, "rgb", None)
            # openpyxl default fill is "00000000" (transparent).
            if rgb and rgb not in {"00000000", None}:
                any_fill = True
    return has_text and not any_bold and not any_fill


def _font_is_cjk(font_name: str | None) -> bool:
    if not font_name:
        return False
    return font_name.strip().lower() in _CJK_FONTS


def _check_workbook_font(wb, issues: list) -> None:
    try:
        normal = wb._named_styles["Normal"]
        font_name = getattr(normal.font, "name", None)
    except Exception:
        font_name = None
    if not _font_is_cjk(font_name):
        issues.append({
            "severity": "warning",
            "code": "WB001",
            "location": "Workbook!Normal",
            "message": (
                f"Workbook default font '{font_name or '<unset>'}' is not a "
                "CJK-capable family. Chinese / Japanese / Korean text may fall "
                "back to system fonts. Set wb._named_styles['Normal'].font = "
                "Font(name='Microsoft YaHei', size=11)."
            ),
        })


def _check_sheet(ws, issues: list) -> None:
    max_col = ws.max_column or 0
    max_row = ws.max_row or 0

    # Wide sheet without freeze panes.
    if max_col >= 8 and max_row >= 20 and not ws.freeze_panes:
        issues.append({
            "severity": "warning",
            "code": "WS001",
            "location": f"{ws.title}",
            "message": (
                f"Wide sheet ({max_col} cols x {max_row} rows) has no "
                "freeze_panes. Set ws.freeze_panes = 'B2' (or similar) so "
                "headers stay visible during scroll."
            ),
        })

    # Header row appears unstyled.
    if max_row >= 2 and max_col >= 2:
        try:
            header_cells = list(ws[1])
            if _looks_unstyled_header(header_cells):
                issues.append({
                    "severity": "warning",
                    "code": "WS002",
                    "location": f"{ws.title}!1:1",
                    "message": (
                        "Header row has no bold and no fill. Apply at least "
                        "Font(bold=True) + a fill color so the table reads as "
                        "structured rather than naked."
                    ),
                })
        except Exception:
            pass


def _scan_numfmts(ws, issues: list, sample_cap: int = 50) -> None:
    seen: set[str] = set()
    sampled = 0
    for row in ws.iter_rows():
        for cell in row:
            fmt = cell.number_format
            if fmt and fmt.lower() not in _KNOWN_NUMFMTS and fmt not in seen:
                seen.add(fmt)
                sampled += 1
                if sampled <= 5:
                    issues.append({
                        "severity": "info",
                        "code": "FMT001",
                        "location": f"{ws.title}!{cell.coordinate}",
                        "message": (
                            f"number_format '{fmt}' is not in the standard "
                            "library (style_kit.NUMFMT). Consider switching "
                            "to a library entry for consistency across files."
                        ),
                    })
                if sampled > sample_cap:
                    return


def _ref_sheet(ref: str) -> str | None:
    m = re.match(r"^\s*(?:'([^']+)'|([^!]+))!", ref or "")
    return (m.group(1) or m.group(2)) if m else None


def _ref_target_cells(ref: str):
    """Parse 'Sheet!A1:Z9' → (sheet_name, c1, r1, c2, r2). Returns None if not a range."""
    m = re.match(
        r"""^\s*(?:'(?P<sheet1>[^']+)'|(?P<sheet2>[^!]+))!
            \$?(?P<c1>[A-Z]+)\$?(?P<r1>\d+)
            (?: : \$?(?P<c2>[A-Z]+)\$?(?P<r2>\d+) )?
            \s*$""",
        ref or "", re.VERBOSE,
    )
    if not m:
        return None
    def _ci(L):
        n = 0
        for c in L.upper():
            n = n * 26 + (ord(c) - 64)
        return n
    sheet = m.group("sheet1") or m.group("sheet2")
    c1 = _ci(m.group("c1")); r1 = int(m.group("r1"))
    c2 = _ci(m.group("c2")) if m.group("c2") else c1
    r2 = int(m.group("r2")) if m.group("r2") else r1
    return sheet, c1, r1, c2, r2


def _has_builder_meta(chart) -> bool:
    return bool(getattr(chart, "_builder_meta", None))


def _gridline_color_ok(gridlines) -> bool:
    """Light-grey dashed line is OK. Default heavy solid (or no spPr) is not."""
    if gridlines is None:
        return True  # absence is fine
    spPr = getattr(gridlines, "spPr", None) or getattr(gridlines, "graphicalProperties", None)
    if spPr is None:
        return False
    ln = getattr(spPr, "ln", None) or getattr(spPr, "line", None)
    if ln is None:
        return False
    fill = getattr(ln, "solidFill", None)
    if fill is None:
        return False
    rgb = getattr(fill, "val", None) or fill if isinstance(fill, str) else None
    # Hex like 'CCCCCC'? Compute brightness.
    if isinstance(rgb, str) and re.fullmatch(r"[0-9A-Fa-f]{6}", rgb):
        brightness = sum(int(rgb[i:i+2], 16) for i in (0, 2, 4)) / 3
        return brightness >= 180
    # Could not introspect → don't flag.
    return True


def _check_chart(ws, chart, idx: int, all_sheets: set, wb, issues: list) -> None:
    series = list(getattr(chart, "series", []) or [])
    # combo charts append overlay charts to chart._charts; include their series too.
    # IMPORTANT: openpyxl puts the chart itself as _charts[0] on reload (self-reference) —
    # we must skip that or every simple chart's series would be counted twice.
    overlay_series = []
    for sub in getattr(chart, "_charts", []) or []:
        if sub is chart:
            continue
        overlay_series.extend(list(getattr(sub, "series", []) or []))
    all_series = series + overlay_series
    n = len(all_series)
    builder = _has_builder_meta(chart)

    if n == 0:
        issues.append({
            "severity": "error", "code": "CH001",
            "location": f"{ws.title}!chart#{idx}",
            "message": "Chart has zero series.",
        })
        return

    if n > 8:
        issues.append({
            "severity": "warning", "code": "CH002",
            "location": f"{ws.title}!chart#{idx}",
            "message": (
                f"Chart has {n} series (> 8). Legend will overflow / repeat. "
                "Consider one series per data row/column, not per data point."
            ),
        })

    # CH003: per-point series antipattern.
    single_cell_count = 0
    for s in all_series:
        val = getattr(s, "val", None)
        nref = getattr(val, "numRef", None) if val else None
        f = getattr(nref, "f", None) if nref else None
        if f and _SINGLE_CELL_RE.match(f):
            single_cell_count += 1
    if single_cell_count == n and n > 1:
        issues.append({
            "severity": "error", "code": "CH003",
            "location": f"{ws.title}!chart#{idx}",
            "message": (
                "Every series's value ref is a single cell. Likely the "
                "per-point series antipattern. Use a multi-cell Reference and "
                "add_data(ref, titles_from_data=True[, from_rows=True])."
            ),
        })

    # CH004: missing categories. Accept refs OR inline literals.
    has_categories = False
    cat_sample = None
    for s in all_series:
        cat = getattr(s, "cat", None)
        if cat is None:
            continue
        sref  = getattr(cat, "strRef",  None)
        nref  = getattr(cat, "numRef",  None)
        slit  = getattr(cat, "strLit",  None)
        nlit  = getattr(cat, "numLit",  None)
        if (sref and getattr(sref, "f", None)) \
                or (nref and getattr(nref, "f", None)) \
                or (slit and (getattr(slit, "pt", None) or getattr(slit, "ptCount", None))) \
                or (nlit and (getattr(nlit, "pt", None) or getattr(nlit, "ptCount", None))):
            has_categories = True
            cat_sample = cat
            break
    if not has_categories:
        issues.append({
            "severity": "error", "code": "CH004",
            "location": f"{ws.title}!chart#{idx}",
            "message": (
                "Chart has no category axis. Legend will show 'Series 1…N'. "
                "Call set_categories() — or build via builders.build_chart()."
            ),
        })

    # CH015: cat numRef points at TEXT cells (X axis renders blank).
    if cat_sample is not None:
        nref = getattr(cat_sample, "numRef", None)
        f = getattr(nref, "f", None) if nref else None
        cache = getattr(nref, "numCache", None) if nref else None
        if f:
            target = _ref_target_cells(f)
            text_in_target = False
            if target and target[0] in all_sheets:
                tws = wb[target[0]]
                _, c1, r1, c2, r2 = target
                for r in range(r1, r2 + 1):
                    for c in range(c1, c2 + 1):
                        v = tws.cell(row=r, column=c).value
                        if isinstance(v, str) and v.strip() != "":
                            text_in_target = True
                            break
                    if text_in_target:
                        break
            cache_pt_count = len(getattr(cache, "pt", []) or []) if cache else 0
            cache_declared = getattr(cache, "ptCount", None)
            cache_declared = getattr(cache_declared, "val", None) if cache_declared else None
            if text_in_target or (cache_declared and cache_pt_count < cache_declared):
                issues.append({
                    "severity": "error", "code": "CH015",
                    "location": f"{ws.title}!chart#{idx}",
                    "message": (
                        "Category axis renders empty: cat is a numRef but the "
                        "target cells contain text (or numCache is unpopulated). "
                        "Excel/WPS/Numbers/Sheet.js will show no X labels. "
                        "Use inline strLit categories — builders.build_chart() "
                        "does this by default."
                    ),
                })

    # CH013 is handled by _scan_chart_xml (post-save XML scan); see below.

    # CH011: axis title set while legend is at bottom — visual collision.
    has_axis_title = False
    for axis_attr in ("x_axis", "y_axis"):
        ax = getattr(chart, axis_attr, None)
        if ax is None:
            continue
        if getattr(ax, "title", None):
            has_axis_title = True
            break
    legend = getattr(chart, "legend", None)
    if has_axis_title and legend and getattr(legend, "position", None) == "b":
        issues.append({
            "severity": "warning", "code": "CH011",
            "location": f"{ws.title}!chart#{idx}",
            "message": (
                "Axis title combined with bottom legend tends to overlap. "
                "Drop the axis title and put units in the chart title (e.g. '($mm)')."
            ),
        })

    # CH012: y-axis numFmt General/empty though data column is $ or %.
    # X axis is a category axis here — no numFmt needed.
    if not builder:
        ax = getattr(chart, "y_axis", None)
        nf = getattr(ax, "numFmt", None) if ax else None
        nf_code = getattr(nf, "formatCode", None) if nf else None
        if ax is not None and nf_code in (None, "", "General") and all_series:
            val = getattr(all_series[0], "val", None)
            nref = getattr(val, "numRef", None) if val else None
            f = getattr(nref, "f", None) if nref else None
            if f:
                target = _ref_target_cells(f)
                if target and target[0] in all_sheets:
                    tws = wb[target[0]]
                    _, c1, r1, c2, r2 = target
                    for r in range(r1, min(r2, r1 + 5) + 1):
                        hit = False
                        for c in range(c1, c2 + 1):
                            fmt = tws.cell(row=r, column=c).number_format or ""
                            if "$" in fmt or "%" in fmt:
                                issues.append({
                                    "severity": "warning", "code": "CH012",
                                    "location": f"{ws.title}!chart#{idx}",
                                    "message": (
                                        f"y-axis numFmt is '{nf_code or '<unset>'}' "
                                        f"but source data uses '{fmt}'. Set "
                                        "ax.number_format or use "
                                        "builders.build_chart(y_number_format=...)."
                                    ),
                                })
                                hit = True
                                break
                        if hit:
                            break

    # CH010: gridlines unstyled / dark default.
    if not builder:  # builder always sets light grey
        gl = getattr(chart, "y_axis", None)
        gl = getattr(gl, "majorGridlines", None) if gl else None
        if gl is not None and not _gridline_color_ok(gl):
            issues.append({
                "severity": "warning", "code": "CH010",
                "location": f"{ws.title}!chart#{idx}",
                "message": (
                    "majorGridlines exist but use the heavy default style. "
                    "Use builders.build_chart() or set "
                    "spPr=GraphicalProperties(ln=LineProperties(solidFill='CCCCCC', prstDash='dash'))."
                ),
            })

    # CH016 / CH017: y axis range hygiene.
    if not builder:
        yax = getattr(chart, "y_axis", None)
        if yax is not None:
            major = getattr(yax, "majorUnit", None)
            scaling = getattr(yax, "scaling", None)
            smin = getattr(scaling, "min", None) if scaling else None
            if major is None:
                issues.append({
                    "severity": "warning", "code": "CH016",
                    "location": f"{ws.title}!chart#{idx}",
                    "message": (
                        "Y axis has no majorUnit; tick labels often render "
                        "as a sparse / unreadable mess. Set y_axis.majorUnit "
                        "or use builders.build_chart()."
                    ),
                })
            # CH017: data clusters near top.
            if smin == 0 or smin is None:
                vals = []
                for s in all_series:
                    cache = getattr(getattr(s, "val", None), "numRef", None)
                    cache = getattr(cache, "numCache", None) if cache else None
                    for pt in (getattr(cache, "pt", []) or []):
                        try:
                            vals.append(float(pt.v))
                        except Exception:
                            pass
                if vals:
                    lo, hi = min(vals), max(vals)
                    if hi > 0 and (hi - lo) / hi < 0.2:
                        issues.append({
                            "severity": "warning", "code": "CH017",
                            "location": f"{ws.title}!chart#{idx}",
                            "message": (
                                "Y axis pinned at 0 but data sits in the top "
                                "1/5 of the range — variations are visually "
                                "flattened. Use y_axis_min='auto_tight' (good "
                                "for pct / ratio / index data)."
                            ),
                        })

    # CH018: bottom legend without overlay=False or explicit layout.
    if legend and getattr(legend, "position", None) == "b":
        overlay = getattr(legend, "overlay", None)
        layout = getattr(legend, "layout", None)
        if overlay is None and layout is None:
            issues.append({
                "severity": "warning", "code": "CH018",
                "location": f"{ws.title}!chart#{idx}",
                "message": (
                    "Bottom legend has neither overlay=False nor an explicit "
                    "layout — Excel will overlap it onto the plot area. "
                    "Use builders.build_chart(legend='bottom')."
                ),
            })

    # CH005: refs to non-existent sheets.
    for s_idx, s in enumerate(all_series):
        for axis_attr in ("val", "cat"):
            axis = getattr(s, axis_attr, None)
            if axis is None:
                continue
            for ref_attr in ("numRef", "strRef"):
                r = getattr(axis, ref_attr, None)
                f = getattr(r, "f", None) if r else None
                if not f:
                    continue
                target = _ref_sheet(f)
                if target and target not in all_sheets:
                    issues.append({
                        "severity": "warning", "code": "CH005",
                        "location": f"{ws.title}!chart#{idx}!ser#{s_idx}",
                        "message": (
                            f"Series ref points to sheet '{target}' which is "
                            f"not in the workbook. Ref: {f}"
                        ),
                    })


def _scan_chart_xml(xlsx_path, issues):
    """Scan xl/charts/*.xml for things openpyxl drops on reload (e.g. <c:style/>)."""
    import zipfile
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            chart_files = [n for n in z.namelist()
                           if n.startswith("xl/charts/chart") and n.endswith(".xml")]
            for cf in chart_files:
                xml = z.read(cf).decode("utf-8", errors="ignore")
                # CH013: <style val="N"/> (or <c:style ...>) with N in 1..48
                m = re.search(r'<(?:c:)?style\s+val="(\d+)"\s*/>', xml)
                if m:
                    n = int(m.group(1))
                    if 1 <= n <= 48:
                        issues.append({
                            "severity": "error", "code": "CH013",
                            "location": cf,
                            "message": (
                                f"chart.style={n} forces an Excel 2007 built-in "
                                "look (heavy gridlines, glossy fills). Remove "
                                "the assignment; let the theme palette take over."
                            ),
                        })
    except Exception:
        pass  # Best-effort; don't fail verification on a packaging quirk.


def _check_chart_sizes(ws, charts, issues):
    """CH014: charts on the same sheet should share a consistent size.

    On reload openpyxl drops chart.width / chart.height back to defaults
    (15 / 7.5 cm); the *real* rendered size lives in the anchor's ext (EMU).
    """
    if len(charts) < 2:
        return
    sizes = []
    for c in charts:
        a = getattr(c, "anchor", None)
        ext = getattr(a, "ext", None) if a else None
        cx = getattr(ext, "cx", None) if ext else None
        cy = getattr(ext, "cy", None) if ext else None
        if cx and cy:
            sizes.append((cx, cy))
    if len(sizes) < 2:
        return
    widths  = [s[0] for s in sizes]
    heights = [s[1] for s in sizes]
    def _spread(xs):
        return (max(xs) - min(xs)) / max(xs) if max(xs) else 0
    if _spread(widths) > 0.30 or _spread(heights) > 0.30:
        issues.append({
            "severity": "warning", "code": "CH014",
            "location": f"{ws.title}",
            "message": (
                f"Charts on this sheet vary by > 30% in size "
                f"(widths={[round(w,1) for w in widths]}, "
                f"heights={[round(h,1) for h in heights]}). Pick one size "
                "preset and pass it to every build_chart(size=...) call."
            ),
        })


def verify_workbook(filename: str) -> dict:
    path = Path(filename)
    if not path.exists():
        return {"status": "error", "error": f"File {filename} does not exist"}

    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"status": "error", "error": "openpyxl is required (pip install openpyxl)"}

    try:
        wb = load_workbook(path, data_only=False)
    except Exception as e:
        return {"status": "error", "error": f"Failed to load workbook: {e}"}

    issues: list = []
    _check_workbook_font(wb, issues)

    all_sheets = set(wb.sheetnames)
    for ws in wb.worksheets:
        _check_sheet(ws, issues)
        _scan_numfmts(ws, issues)
        charts = list(getattr(ws, "_charts", []) or [])
        for c_idx, chart in enumerate(charts):
            _check_chart(ws, chart, c_idx, all_sheets, wb, issues)
        _check_chart_sizes(ws, charts, issues)

    _scan_chart_xml(path, issues)

    summary = {
        "errors": sum(1 for i in issues if i["severity"] == "error"),
        "warnings": sum(1 for i in issues if i["severity"] == "warning"),
        "infos": sum(1 for i in issues if i["severity"] == "info"),
    }
    status = "issues_found" if (summary["errors"] or summary["warnings"]) else "success"
    return {"status": status, "summary": summary, "issues": issues[:100]}


def main() -> None:
    if len(sys.argv) < 2:
        print(
            "Usage: python verify_workbook.py <xlsx_path>\n\n"
            "Structural lint pass: chart series, categories, freeze panes,\n"
            "header styling, CJK font, number formats."
        )
        sys.exit(1)
    print(json.dumps(verify_workbook(sys.argv[1]), indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
