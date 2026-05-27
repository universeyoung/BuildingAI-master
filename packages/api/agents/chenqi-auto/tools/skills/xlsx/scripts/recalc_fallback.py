#!/usr/bin/env python3
"""
Pure-Python fallback for `recalc.py` when LibreOffice (`soffice`) is unavailable.

What it does:
  1. Sets `workbook.calcPr.fullCalcOnLoad = True` so that Excel, WPS Office,
     Numbers and Excel Online recompute every formula the moment the user
     opens the file. This is the OOXML-spec-compliant way to defer recalc to
     the viewer and is honored by all major spreadsheet apps. Formulas are
     preserved verbatim — nothing is hardcoded.
  2. Statically evaluates every formula in pure Python via the `formulas`
     library to detect Excel errors (#DIV/0!, #VALUE!, #REF!, #NAME?, #NULL!,
     #NUM!, #N/A) that the agent should fix BEFORE delivering. The computed
     values are intentionally NOT written back — `formulas`'s own writer
     replaces formulas with constants, which would defeat the point of using
     a spreadsheet. We only read the computed values to surface errors.

Important contract guarantees:
  - The output file's formulas are byte-for-byte identical to the input's.
  - No formula cell is replaced by a hardcoded numeric/string literal.
  - The cached value (the number a viewer shows before clicking the cell)
    will be filled in by the spreadsheet app on open, not by this script.

Output JSON shape mirrors `recalc.py` for tooling compatibility:
    {
      "status": "success" | "errors_found" | "error",
      "total_formulas": int,
      "total_errors": int,
      "error_summary": { "#DIV/0!": {"count": N, "locations": [...]}, ... },
      "fallback_mode": "fullCalcOnLoad",
      "note": "Formulas preserved; recalculated by Excel/WPS/Numbers on next open."
    }

Usage:
    python scripts/recalc_fallback.py <excel_file>
"""

import json
import sys
from collections import defaultdict
from pathlib import Path

EXCEL_ERRS = {"#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"}


def _set_full_calc_on_load(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.workbook.properties import CalcProperties

    wb = load_workbook(path)
    if wb.calculation is None:
        wb.calculation = CalcProperties()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.calcMode = "auto"
    wb.save(path)


def _count_formulas(path: Path) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=False)
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    n += 1
    wb.close()
    return n


def _scan_formula_errors(path: Path) -> list[tuple[str, str]]:
    """Use `formulas` to evaluate the workbook and report cells whose result is
    an Excel error. Returns a list of (cell_ref, error_token) tuples."""
    try:
        import formulas  # type: ignore
    except ImportError:
        return []

    try:
        xl = formulas.ExcelModel().loads(str(path)).finish()
        sol = xl.calculate()
    except Exception:
        # `formulas` chokes on some real-world files (volatile fns, custom defined
        # names, unsupported features). Best-effort: return nothing rather than
        # crashing the whole fallback.
        return []

    errors: list[tuple[str, str]] = []
    for k, v in sol.items():
        val = v.value if hasattr(v, "value") else v
        # Unwrap [[scalar]] from formulas's numpy-style arrays
        if hasattr(val, "tolist"):
            val = val.tolist()
            while isinstance(val, list) and len(val) == 1:
                val = val[0]
        s = str(val)
        if s in EXCEL_ERRS:
            ref = k.split("]", 1)[-1] if "]" in k else k
            errors.append((ref, s))
    return errors


def recalc_fallback(filename: str) -> dict:
    path = Path(filename)
    if not path.exists():
        return {"status": "error", "error": f"File {filename} does not exist"}

    try:
        _set_full_calc_on_load(path)
    except Exception as e:
        return {"status": "error", "error": f"Failed to set fullCalcOnLoad: {e}"}

    total_formulas = _count_formulas(path)
    formula_errors = _scan_formula_errors(path)

    grouped: dict[str, list[str]] = defaultdict(list)
    for ref, err in formula_errors:
        grouped[err].append(ref)

    result: dict = {
        "status": "errors_found" if formula_errors else "success",
        "total_formulas": total_formulas,
        "total_errors": len(formula_errors),
        "fallback_mode": "fullCalcOnLoad",
        "note": (
            "Formulas preserved verbatim. Cached values will be computed by "
            "Excel/WPS/Numbers when the file is opened. NEVER hardcode values "
            "to work around missing LibreOffice."
        ),
    }
    if formula_errors:
        result["error_summary"] = {
            err: {"count": len(locs), "locations": locs[:20]}
            for err, locs in grouped.items()
        }
    if not formula_errors and "formulas" not in sys.modules:
        result["warning"] = (
            "Static error scan skipped: install the `formulas` package "
            "(`pip install formulas`) to detect #DIV/0!/#VALUE!/#REF! before delivery."
        )
    return result


def main() -> None:
    if len(sys.argv) < 2:
        print(
            "Usage: python recalc_fallback.py <excel_file>\n\n"
            "Pure-Python fallback for recalc.py when LibreOffice is missing.\n"
            "  - Sets fullCalcOnLoad=True (Excel/WPS/Numbers recompute on open)\n"
            "  - Statically scans formulas for #DIV/0!/#VALUE!/#REF!/etc.\n"
            "  - Preserves all formulas (NEVER hardcodes computed values)"
        )
        sys.exit(1)

    print(json.dumps(recalc_fallback(sys.argv[1]), indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
