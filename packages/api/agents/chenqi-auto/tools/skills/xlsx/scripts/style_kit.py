#!/usr/bin/env python3
"""
Reusable styling primitives for openpyxl-built workbooks.

Why this exists
---------------
SKILL.md used to describe number_format / color rules in prose. Models then
re-derived the exact strings each time, producing 6 inconsistent variants of
the "same" format across files. This module turns those decisions into
named constants and two helper functions.

Public surface
--------------
NUMFMT          dict[str, str]    common number_format strings
PALETTES        dict[str, list[str]]  ordered series-color palettes
THEMES          dict[str, dict]   theme bundles (font, header, accent, palette)
apply_table_style(ws, *, header_row, data_rows, ...)
apply_chart_style(chart, *, theme=..., cjk_font=..., legend_position=...)

Nothing here knows about populate_chart_cache.py or verify_workbook.py — the
three modules are intentionally orthogonal.
"""

from __future__ import annotations

# ---------------------------------------------------------------------------
# NUMFMT — copy/paste these instead of inventing new strings.
# Naming convention: <unit>[_<precision>][_neg]
#   - default forms include negative-paren + zero-as-dash for finance use
#   - simple_* variants omit those rules (use when raw display is preferred)
# ---------------------------------------------------------------------------
NUMFMT: dict[str, str] = {
    # Integers & decimals (with negative-paren + zero-dash)
    "int":           '#,##0;[Red](#,##0);"-"',
    "dec_1":         '#,##0.0;[Red](#,##0.0);"-"',
    "dec_2":         '#,##0.00;[Red](#,##0.00);"-"',

    # Currency (USD; for other currencies, swap the "$" prefix)
    "dollar":        '$#,##0;[Red]($#,##0);"-"',
    "dollar_2":      '$#,##0.00;[Red]($#,##0.00);"-"',

    # Percentages
    "pct_0":         '0%;[Red](0%);"-"',
    "pct_1":         '0.0%;[Red](0.0%);"-"',
    "pct_2":         '0.00%;[Red](0.00%);"-"',

    # Valuation multiples (e.g. "12.5x")
    "multiple":      '0.0"x"',

    # Dates
    "date":          'yyyy-mm-dd',
    "date_us":       'mm/dd/yyyy',
    "date_q":        '"Q"q yyyy',
    "date_my":       'mmm yyyy',

    # Misc
    "text":          '@',
    "general":       'General',
    # "Year as text" — prevents "2024" being thousand-separated to "2,024"
    "year":          '0',

    # Simple variants (no neg-paren, no zero-dash)
    "simple_int":    '#,##0',
    "simple_dec_2":  '#,##0.00',
    "simple_dollar": '$#,##0',
    "simple_pct_1":  '0.0%',
}


# ---------------------------------------------------------------------------
# PALETTES — series colors, hex (no leading "#").
# ---------------------------------------------------------------------------
PALETTES: dict[str, list[str]] = {
    "modern_finance": [
        "1F3864", "ED7D31", "70AD47", "A5A5A5",
        "5B9BD5", "FFC000", "264478", "9E480E",
    ],
    "minimal_grey": [
        "2F2F2F", "595959", "808080", "A6A6A6",
        "BFBFBF", "D9D9D9",
    ],
    "data_dense": [
        "1F77B4", "FF7F0E", "2CA02C", "D62728",
        "9467BD", "8C564B", "E377C2", "7F7F7F",
    ],
    # Deep claret + gold + slate; restrained, authoritative.
    "executive_report": [
        "6B1F2A", "C9A227", "3E5C76", "748CAB",
        "8C5E58", "B6A07A", "4F6D7A", "C44536",
    ],
    # High-contrast modern blue/orange/teal/magenta; energetic.
    "vibrant_marketing": [
        "0066CC", "FF6B35", "00B894", "E84393",
        "FDCB6E", "6C5CE7", "00A8E8", "F368E0",
    ],
    # Indigo / cyan / lime; cool SaaS dashboard look.
    "tech_product": [
        "4C3A99", "00B4D8", "90BE6D", "F4A261",
        "577590", "E76F51", "43AA8B", "9D4EDD",
    ],
    # Pure neutrals, max contrast; ink-saving for printers.
    "print_friendly": [
        "000000", "404040", "707070", "9A9A9A",
        "BDBDBD", "DCDCDC",
    ],
}


# ---------------------------------------------------------------------------
# THEMES — bundled visual choices. Keys: font, header_*, accent, palette.
# ---------------------------------------------------------------------------
THEMES: dict[str, dict] = {
    "modern_finance": {
        "font_name":     "Microsoft YaHei",  # CJK-capable + clean Latin
        "font_size":     11,
        "header_fill":   "1F3864",           # navy
        "header_font":   "FFFFFF",           # white
        "header_bold":   True,
        "accent":        "ED7D31",           # orange highlight
        "row_band":      "F2F2F2",           # very light grey
        "border":        "BFBFBF",
        "palette":       "modern_finance",
    },
    "minimal_grey": {
        "font_name":     "Microsoft YaHei",
        "font_size":     11,
        "header_fill":   "2F2F2F",
        "header_font":   "FFFFFF",
        "header_bold":   True,
        "accent":        "595959",
        "row_band":      "F7F7F7",
        "border":        "D9D9D9",
        "palette":       "minimal_grey",
    },
    "data_dense": {
        "font_name":     "Microsoft YaHei",
        "font_size":     10,                 # smaller for dense tables
        "header_fill":   "264478",
        "header_font":   "FFFFFF",
        "header_bold":   True,
        "accent":        "1F77B4",
        "row_band":      "FAFAFA",
        "border":        "E5E5E5",
        "palette":       "data_dense",
    },
    "executive_report": {
        "font_name":     "Microsoft YaHei",
        "font_size":     11,
        "header_fill":   "6B1F2A",           # deep claret
        "header_font":   "FFFFFF",
        "header_bold":   True,
        "accent":        "C9A227",           # antique gold
        "row_band":      "F7F2EC",           # warm cream
        "border":        "C9B8A8",
        "palette":       "executive_report",
    },
    "vibrant_marketing": {
        "font_name":     "Microsoft YaHei",
        "font_size":     11,
        "header_fill":   "0066CC",           # vivid blue
        "header_font":   "FFFFFF",
        "header_bold":   True,
        "accent":        "FF6B35",           # punchy orange
        "row_band":      "EEF6FC",           # very light blue
        "border":        "C7D8E8",
        "palette":       "vibrant_marketing",
    },
    "tech_product": {
        "font_name":     "Microsoft YaHei",
        "font_size":     11,
        "header_fill":   "2D2A4A",           # dark indigo
        "header_font":   "FFFFFF",
        "header_bold":   True,
        "accent":        "00B4D8",           # cyan highlight
        "row_band":      "F4F4F8",
        "border":        "D8D8E2",
        "palette":       "tech_product",
    },
    "print_friendly": {
        "font_name":     "Microsoft YaHei",
        "font_size":     11,
        "header_fill":   "000000",           # pure black
        "header_font":   "FFFFFF",
        "header_bold":   True,
        "accent":        "404040",
        "row_band":      "F2F2F2",           # very light grey
        "border":        "808080",
        "palette":       "print_friendly",
    },
}


# ---------------------------------------------------------------------------
# apply_table_style — minimum-viable table polish in one call.
# ---------------------------------------------------------------------------
def apply_table_style(
    ws,
    *,
    header_row: int = 1,
    data_rows: tuple[int, int] | None = None,
    total_row: int | None = None,
    theme: str = "modern_finance",
    freeze: str = "none",
    band_rows: bool = True,
) -> None:
    """Apply a baseline visual treatment to a sheet.

    Args:
      ws:          openpyxl Worksheet
      header_row:  1-based row index of the header
      data_rows:   (start, end) inclusive; defaults to (header_row+1, ws.max_row)
      total_row:   optional 1-based row index for the totals row
      theme:       key into THEMES
      freeze:      'none' | 'rows' | 'cols' | 'both'
                   - 'rows': freeze just below header (header always visible)
                   - 'cols': freeze right of col A (label column always visible)
                   - 'both': freeze both axes (anchor B<header+1>)
                   - 'none': no freeze (small tables don't need it)
      band_rows:   if True, alternate-row fill for readability
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if theme not in THEMES:
        raise ValueError(f"Unknown theme '{theme}'. Choose from: {list(THEMES)}")
    t = THEMES[theme]

    base_font = Font(name=t["font_name"], size=t["font_size"])
    header_font = Font(
        name=t["font_name"], size=t["font_size"],
        bold=t["header_bold"], color=t["header_font"],
    )
    header_fill = PatternFill("solid", start_color=t["header_fill"])
    band_fill = PatternFill("solid", start_color=t["row_band"])
    thin = Side(style="thin", color=t["border"])
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    max_col = ws.max_column or 1
    max_row = ws.max_row or header_row
    if data_rows is None:
        data_rows = (header_row + 1, max_row if total_row is None else total_row - 1)
    d_start, d_end = data_rows

    # Header row
    for c in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = cell_border

    # Data rows.
    # NOTE: openpyxl always returns a non-None Font on a cell (defaults to
    # Calibri 11) and the workbook-default 'Normal' style does NOT propagate
    # once a cell gets any other style attribute (e.g. number_format). So the
    # only reliable way to ensure CJK glyphs render is to *unconditionally*
    # stamp the theme font onto every data cell.
    for r in range(d_start, d_end + 1):
        fill = band_fill if (band_rows and (r - d_start) % 2 == 1) else None
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = base_font
            if fill is not None and (cell.fill is None or
                                     getattr(cell.fill.fgColor, "rgb", None) in
                                     {"00000000", None}):
                cell.fill = fill
            cell.border = cell_border

    # Total row
    if total_row:
        total_font = Font(name=t["font_name"], size=t["font_size"], bold=True)
        top_double = Side(style="double", color=t["header_fill"])
        for c in range(1, max_col + 1):
            cell = ws.cell(row=total_row, column=c)
            cell.font = total_font
            cell.border = Border(top=top_double, bottom=thin, left=thin, right=thin)

    # Freeze panes — granular: rows / cols / both / none.
    if freeze not in ("none", "rows", "cols", "both"):
        raise ValueError(
            f"Unknown freeze mode '{freeze}'. "
            f"Choose from: 'none', 'rows', 'cols', 'both'"
        )
    if freeze == "rows":
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    elif freeze == "cols":
        ws.freeze_panes = ws.cell(row=1, column=2).coordinate
    elif freeze == "both":
        ws.freeze_panes = ws.cell(row=header_row + 1, column=2).coordinate


# ---------------------------------------------------------------------------
# apply_chart_style — set CJK-safe fonts on chart text, plus theme polish.
# ---------------------------------------------------------------------------
def apply_chart_style(
    chart,
    *,
    theme: str = "modern_finance",
    cjk_font: str | None = None,
    legend_position: str | None = "b",
) -> None:
    """Apply baseline chart styling: CJK font on title/legend/axes, palette.

    The CJK font is stamped onto title, legend, and axis text via the
    DrawingML `defRPr.latin / ea` attributes — without this, charts on a
    machine missing Calibri's CJK fallback render boxes for Chinese text.
    """
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import (
        Paragraph, ParagraphProperties, CharacterProperties, RichTextProperties,
        Font as DrawingFont,
    )

    if theme not in THEMES:
        raise ValueError(f"Unknown theme '{theme}'. Choose from: {list(THEMES)}")
    t = THEMES[theme]
    font_name = cjk_font or t["font_name"]

    # Soft deprecation: charts not created via builders.build_chart skip the
    # axis-scaling / literal-cats / dashed-gridlines defaults that
    # verify_workbook (CH010 / CH015 / CH016 / CH017) checks for. We don't
    # break — builders calls us internally and sets _builder_meta first.
    if not getattr(chart, "_builder_meta", None):
        import warnings
        warnings.warn(
            "apply_chart_style was called on a chart not created by "
            "builders.build_chart; verify_workbook will likely flag CH010 / "
            "CH015 / CH016. Prefer build_chart() for new charts.",
            DeprecationWarning, stacklevel=2,
        )

    def _rich(font_size_100: int = 1100):
        cp = CharacterProperties(
            sz=font_size_100,
            latin=DrawingFont(typeface=font_name),
            ea=DrawingFont(typeface=font_name),
        )
        pp = ParagraphProperties(defRPr=cp)
        # NOTE: Paragraph(pPr=...) auto-emits an empty <a:r><a:t/></a:r> Run.
        # Pass r=[] to suppress it — otherwise some Excel builds render the
        # placeholder text as the literal word "None" inside the axis txPr.
        return RichText(bodyPr=RichTextProperties(),
                        p=[Paragraph(pPr=pp, r=[])])

    # Title
    if getattr(chart, "title", None) is not None:
        try:
            chart.title.tx.rich.p[0].pPr = _rich(1400).p[0].pPr
        except Exception:
            pass

    # Axes
    for axis_attr in ("x_axis", "y_axis"):
        ax = getattr(chart, axis_attr, None)
        if ax is not None:
            try:
                ax.txPr = _rich(1000)
            except Exception:
                pass

    # Legend
    if legend_position is not None:
        try:
            from openpyxl.chart.legend import Legend
            if chart.legend is None:
                chart.legend = Legend()
            chart.legend.position = legend_position
            chart.legend.txPr = _rich(1000)
        except Exception:
            pass

    # Apply palette colors series-by-series
    palette = PALETTES.get(t["palette"], [])
    if palette:
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.fill import ColorChoice
        for i, ser in enumerate(getattr(chart, "series", []) or []):
            color = palette[i % len(palette)]
            try:
                ser.graphicalProperties = GraphicalProperties(solidFill=color)
            except Exception:
                pass


# ---------------------------------------------------------------------------
# Polish helpers — operate on the handle dict returned by build_table.
#
# All helpers below are *additive*: call them AFTER build_table to layer
# group coloring, conditional highlights, data bars, color scales, icon sets,
# and column separators on top of the baseline theme. They never re-write
# data; they only set fills / fonts / borders / conditional formatting.
#
# The handle `t` is:
#   {'header_row': int, 'data_rows': (start, end), 'total_row': int|None,
#    'columns': {name: 'A'}, 'sheet': str, ...}
# ---------------------------------------------------------------------------

# Light tints suitable for column-group backgrounds. Pair freely; they share
# luminance so a 3-group layout reads as "sections", not "warnings".
TINTS: dict[str, str] = {
    "mint":     "E6F4EA",
    "sky":      "E3F0FB",
    "peach":    "FCEBDC",
    "lavender": "EFE7F7",
    "butter":   "FFF8DC",
    "rose":     "FBE5E8",
    "slate":    "ECEEF1",
    "sand":     "F4EFE3",
}

# Named cell-emphasis styles for highlight_cells. Each entry is a dict of
# kwargs forwarded to openpyxl Font / PatternFill (handled in the helper).
CELL_STYLES: dict[str, dict] = {
    "red_text":    {"font_color": "C00000", "bold": True},
    "green_text":  {"font_color": "1E7E34", "bold": True},
    "amber_text":  {"font_color": "B7791F", "bold": True},
    "red_fill":    {"fill": "F8CBAD", "font_color": "8B0000", "bold": True},
    "green_fill":  {"fill": "C6EFCE", "font_color": "1E5631", "bold": True},
    "amber_fill":  {"fill": "FFEB9C", "font_color": "7F5F00", "bold": True},
    "bold":        {"bold": True},
}

# Data-bar / color-scale named palettes. Hex without leading "#".
_DATA_BAR_COLORS: dict[str, str] = {
    "sky": "5B9BD5", "mint": "70AD47", "peach": "ED7D31",
    "rose": "C00000", "lavender": "7030A0", "slate": "595959",
}
_COLOR_SCALES: dict[str, tuple[str, str, str | None]] = {
    "red_white_green":  ("F8696B", "FFFFFF", "63BE7B"),
    "green_white_red":  ("63BE7B", "FFFFFF", "F8696B"),
    "white_blue":       ("FFFFFF", None,     "5B9BD5"),
    "white_red":        ("FFFFFF", None,     "F8696B"),
}
_ICON_SETS = {
    "3_arrows":   "3Arrows",
    "3_traffic":  "3TrafficLights1",
    "3_symbols":  "3Symbols",
    "3_flags":    "3Flags",
    "5_arrows":   "5Arrows",
    "5_quarters": "5Quarters",
}


def _resolve_columns(t: dict, columns) -> list[tuple[str, str]]:
    """Map column names → (name, letter). Raises on unknown name."""
    out = []
    cols = t["columns"]
    for name in columns:
        if name not in cols:
            raise ValueError(
                f"Unknown column '{name}'. Available: {list(cols)}"
            )
        out.append((name, cols[name]))
    return out


def color_column_groups(ws, t: dict, groups: list[dict]) -> None:
    """Paint contiguous column groups with named tints.

    Args:
      groups: list of {'columns': [name, ...], 'tint': key_in_TINTS}.
              Columns within a group should be contiguous; non-contiguous
              members are still painted but won't read as one block.

    Re-applies row banding *inside* each group (slightly darker than the
    tint) so banded readability survives the recolor.
    """
    from openpyxl.styles import PatternFill
    d_start, d_end = t["data_rows"]
    for g in groups:
        if g["tint"] not in TINTS:
            raise ValueError(f"Unknown tint '{g['tint']}'. Choose from: {list(TINTS)}")
        base = TINTS[g["tint"]]
        band = _darken(base, 0.06)
        base_fill = PatternFill("solid", start_color=base)
        band_fill = PatternFill("solid", start_color=band)
        for _, letter in _resolve_columns(t, g["columns"]):
            col_idx = _col_index(letter)
            for r in range(d_start, d_end + 1):
                cell = ws.cell(row=r, column=col_idx)
                cell.fill = band_fill if (r - d_start) % 2 == 1 else base_fill


_HL_OPERATORS: dict[str, str] = {
    "lt":      "lessThan",
    "le":      "lessThanOrEqual",
    "gt":      "greaterThan",
    "ge":      "greaterThanOrEqual",
    "eq":      "equal",
    "ne":      "notEqual",
    "between": "between",
}


def highlight_cells(ws, t: dict, rules: list[dict]) -> None:
    """Install Excel-native conditional-formatting rules so the highlight
    follows the data — if the user later edits a value or the underlying
    formula recalculates, Excel re-evaluates the rule on the fly.

    Args:
      rules: list of dicts with keys:
        - column:  column name (must exist in t['columns'])
        - when:    'lt' | 'le' | 'gt' | 'ge' | 'eq' | 'ne' | 'between'
        - value:   number/str for lt/le/gt/ge/eq/ne;
                   (lo, hi) tuple for 'between'
        - style:   key in CELL_STYLES

    Each rule becomes one openpyxl `CellIsRule` over the column's data
    range. Works identically for literal values and formula cells.
    """
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Font, PatternFill
    d_start, d_end = t["data_rows"]

    for rule in rules:
        if rule["when"] not in _HL_OPERATORS:
            raise ValueError(
                f"Unknown 'when': {rule['when']}. "
                f"Choose from: {list(_HL_OPERATORS)}"
            )
        if rule["style"] not in CELL_STYLES:
            raise ValueError(
                f"Unknown style '{rule['style']}'. "
                f"Choose from: {list(CELL_STYLES)}"
            )
        op_name = _HL_OPERATORS[rule["when"]]
        spec = CELL_STYLES[rule["style"]]

        # Build the formula list expected by CellIsRule.
        if rule["when"] == "between":
            lo, hi = rule["value"]
            formulas_arg = [_hl_literal(lo), _hl_literal(hi)]
        else:
            formulas_arg = [_hl_literal(rule["value"])]

        # Build the dxf (font + fill) from the named CELL_STYLES entry.
        font_kwargs = {"bold": spec.get("bold", False)}
        if "font_color" in spec:
            font_kwargs["color"] = spec["font_color"]
        dxf_font = Font(**font_kwargs)
        dxf_fill = (
            PatternFill("solid",
                        start_color=spec["fill"], end_color=spec["fill"])
            if "fill" in spec else None
        )

        col_letter = _resolve_columns(t, [rule["column"]])[0][1]
        rng = f"{col_letter}{d_start}:{col_letter}{d_end}"
        cf_rule = CellIsRule(
            operator=op_name, formula=formulas_arg,
            font=dxf_font, fill=dxf_fill, stopIfTrue=False,
        )
        ws.conditional_formatting.add(rng, cf_rule)


def _hl_literal(v) -> str:
    """Render a Python value as an Excel CF-rule formula literal.
    Numbers stay bare; strings are wrapped in double quotes."""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        return repr(v)
    return f'"{str(v)}"'


def add_data_bars(ws, t: dict, *, columns, color: str = "sky",
                  show_value: bool = True) -> None:
    """Add native Excel data bars over the given columns' data range.

    Args:
      columns:    list of column names (must exist in t['columns'])
      color:      key in _DATA_BAR_COLORS (sky/mint/peach/rose/lavender/slate)
      show_value: if False, hides the cell value (bar only)
    """
    from openpyxl.formatting.rule import DataBarRule
    if color not in _DATA_BAR_COLORS:
        raise ValueError(f"Unknown data-bar color '{color}'. "
                         f"Choose from: {list(_DATA_BAR_COLORS)}")
    bar_color = _DATA_BAR_COLORS[color]
    d_start, d_end = t["data_rows"]
    for _, letter in _resolve_columns(t, columns):
        rng = f"{letter}{d_start}:{letter}{d_end}"
        rule = DataBarRule(
            start_type="min", end_type="max",
            color=bar_color, showValue=show_value,
        )
        ws.conditional_formatting.add(rng, rule)


def add_color_scale(ws, t: dict, *, columns, scale: str = "red_white_green") -> None:
    """Add a 2- or 3-color heatmap scale over the given columns' data range.

    Built-in scales: red_white_green, green_white_red, white_blue, white_red.
    """
    from openpyxl.formatting.rule import ColorScaleRule
    if scale not in _COLOR_SCALES:
        raise ValueError(f"Unknown color scale '{scale}'. "
                         f"Choose from: {list(_COLOR_SCALES)}")
    lo, mid, hi = _COLOR_SCALES[scale]
    d_start, d_end = t["data_rows"]
    for _, letter in _resolve_columns(t, columns):
        rng = f"{letter}{d_start}:{letter}{d_end}"
        if mid is None:
            rule = ColorScaleRule(
                start_type="min", start_color=lo,
                end_type="max",   end_color=hi,
            )
        else:
            rule = ColorScaleRule(
                start_type="min",         start_color=lo,
                mid_type="percentile",    mid_value=50, mid_color=mid,
                end_type="max",           end_color=hi,
            )
        ws.conditional_formatting.add(rng, rule)


def add_icon_set(ws, t: dict, *, columns, icon_set: str = "3_arrows") -> None:
    """Add Excel icon-set conditional formatting over the given columns.

    Sets: 3_arrows, 3_traffic, 3_symbols, 3_flags, 5_arrows, 5_quarters.
    """
    from openpyxl.formatting.rule import IconSetRule
    if icon_set not in _ICON_SETS:
        raise ValueError(f"Unknown icon set '{icon_set}'. "
                         f"Choose from: {list(_ICON_SETS)}")
    name = _ICON_SETS[icon_set]
    d_start, d_end = t["data_rows"]
    for _, letter in _resolve_columns(t, columns):
        rng = f"{letter}{d_start}:{letter}{d_end}"
        rule = IconSetRule(name, "percent",
                           [0, 33, 67] if name.startswith("3") else
                           [0, 20, 40, 60, 80])
        ws.conditional_formatting.add(rng, rule)


def add_column_separators(ws, t: dict, *, after_columns: list[str],
                          weight: str = "medium") -> None:
    """Draw a vertical separator on the right edge of each named column,
    spanning from the title row (if any) through the total row (if any).

    Use to visually divide business sections (e.g. "Margins | Cloud | Mix").
    """
    from openpyxl.styles import Border, Side
    side = Side(style=weight, color=THEMES["modern_finance"]["header_fill"])
    r0 = t.get("title_row") or t["header_row"]
    r1 = t.get("total_row") or t["data_rows"][1]
    for _, letter in _resolve_columns(t, after_columns):
        col_idx = _col_index(letter)
        for r in range(r0, r1 + 1):
            cell = ws.cell(row=r, column=col_idx)
            existing = cell.border
            cell.border = Border(
                left=existing.left, top=existing.top, bottom=existing.bottom,
                right=side,
            )


def _col_index(letter: str) -> int:
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def _darken(hex_color: str, ratio: float) -> str:
    """Darken a hex color by `ratio` (0..1). Used to derive banding from a tint."""
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    r = max(0, int(r * (1 - ratio)))
    g = max(0, int(g * (1 - ratio)))
    b = max(0, int(b * (1 - ratio)))
    return f"{r:02X}{g:02X}{b:02X}"


__all__ = [
    "NUMFMT",
    "PALETTES",
    "THEMES",
    "TINTS",
    "CELL_STYLES",
    "apply_table_style",
    "apply_chart_style",
    "color_column_groups",
    "highlight_cells",
    "add_data_bars",
    "add_color_scale",
    "add_icon_set",
    "add_column_separators",
]
